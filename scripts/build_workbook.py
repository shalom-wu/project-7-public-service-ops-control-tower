"""Build excel/public_service_ops_control_tower.xlsx — structure, data,
formulas and formatting (openpyxl). Real pivot tables, pivot charts and the
recalculation pass are added afterwards by finalize_workbook.py (Excel COM).

All metrics in the workbook are live Excel formulas over the Cleaned_Data
table — nothing is precomputed in Python except row-label lists (top agencies
/ complaint types), which are design choices, not calculations.
"""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "excel" / "public_service_ops_control_tower.xlsx"

# ---------------------------------------------------------------------------
# style constants
# ---------------------------------------------------------------------------
DARK = "26343F"
TEAL = "1F7A8C"
LIGHT = "F2F6F8"
AMBER = "FDEBD0"
RED_TXT = "C0392B"
FONT = "Calibri"

F_TITLE = Font(name=FONT, size=16, bold=True, color=DARK)
F_H2 = Font(name=FONT, size=12, bold=True, color=DARK)
F_HDR = Font(name=FONT, size=10, bold=True, color="FFFFFF")
F_BODY = Font(name=FONT, size=10, color="333333")
F_INPUT = Font(name=FONT, size=10, color="0000FF")          # blue = input
F_FORMULA = Font(name=FONT, size=10, color="000000")        # black = formula
F_LINK = Font(name=FONT, size=10, color="008000")           # green = x-sheet
F_NOTE = Font(name=FONT, size=9, italic=True, color="808080")

FILL_HDR = PatternFill("solid", start_color=DARK)
FILL_LIGHT = PatternFill("solid", start_color=LIGHT)
FILL_AMBER = PatternFill("solid", start_color=AMBER)
FILL_KPI = PatternFill("solid", start_color="EAF2F4")

THIN = Side(style="thin", color="D5DBDF")
B_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row, first_col, last_col):
    for c in range(first_col, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = F_HDR
        cell.fill = FILL_HDR
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = B_ALL


def note(ws, cell, text):
    ws[cell] = text
    ws[cell].font = F_NOTE


def title(ws, text, sub=None):
    ws["A1"] = text
    ws["A1"].font = F_TITLE
    if sub:
        ws["A2"] = sub
        ws["A2"].font = F_NOTE


# ---------------------------------------------------------------------------
# load prepared data
# ---------------------------------------------------------------------------
sample = pd.read_csv(ROOT / "data/processed/excel_sample_100k.csv", dtype=str)
raw_preview = pd.read_csv(ROOT / "data/processed/raw_preview_2k.csv", dtype=str)
sla = pd.read_csv(ROOT / "data/assumptions/sla_targets.csv")
import re
meta = (ROOT / "data/processed/extract_metadata.txt").read_text()
TOTAL_ROWS = int(re.search(r"Total after de-dup:\s*([\d,]+)", meta)
                 .group(1).replace(",", ""))

for col in ("created_date", "closed_date"):
    sample[col] = pd.to_datetime(sample[col])

N = len(sample)
LAST = N + 1  # last data row in Cleaned_Data

top_agencies = sample["agency"].value_counts().head(12).index.tolist()
top_complaints = sample["complaint_type"].value_counts().head(15).index.tolist()
boroughs = ["Brooklyn", "Queens", "Manhattan", "Bronx", "Staten Island"]

wb = Workbook()

# ===========================================================================
# 1. README
# ===========================================================================
ws = wb.active
ws.title = "README"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
for col, w in (("B", 28), ("C", 95)):
    ws.column_dimensions[col].width = w
ws["B2"] = "Public Service Operations — SLA Control Tower"
ws["B2"].font = Font(name=FONT, size=18, bold=True, color=DARK)
ws["B3"] = ("Weekly/monthly operations monitor for NYC 311 service requests: demand, closure performance, "
            "backlog, aging, SLA risk and operational priorities.")
ws["B3"].font = F_BODY
rows = [
    ("", ""),
    ("WHAT THIS IS", "An Excel control tower built on public NYC Open Data (311 Service Requests, dataset erm2-nwe9). "
     "It is public city service-request data, NOT private company operations data. Staffing and internal agency "
     "capacity are not in the dataset and are not claimed anywhere in this workbook."),
    ("DATA BASIS", "A 100,000-row random sample (seed 42) of the full 1.07M-row extract: every request created "
     "2026-04-01 to 2026-06-30 plus all older still-open requests back to 2025-01-01. Rates and averages are "
     "unbiased on a random sample; volume KPIs are labeled 'sample' and a scaling factor on the Assumptions tab "
     "grosses them up. Full-population numbers come from the SQL layer (sql/) and Power BI."),
    ("HOW TO USE", "1) Review Assumptions (blue cells are editable inputs — SLA targets, aging buckets, priority weights). "
     "2) Pivot_Analysis for demand cuts. 3) SLA_Analysis and Backlog_Analysis for diagnostics. "
     "4) Priority_Model ranks complaint-type x borough hotspots. 5) Action_List is the takeaway view."),
    ("TAB GUIDE", "Raw_Data_Sample — 2,000 rows exactly as pulled  |  Cleaned_Data — analysis table + derived fields  |  "
     "Data_Dictionary — every field defined  |  Assumptions — all editable inputs  |  Pivot_Analysis — pivot tables & charts  |  "
     "SLA_Analysis — met/missed by agency, type, borough, month  |  Backlog_Analysis — open requests & aging  |  "
     "Priority_Model — transparent weighted scoring  |  Action_List — top priorities in plain English"),
    ("COLOR KEY", "Blue text = editable input/assumption.  Black = formula.  Green = value linked from another tab.  "
     "Amber fill = key assumption. Conditional formatting flags SLA misses and aged backlog."),
    ("SLA CAVEAT", "NYC publishes no per-request-type SLA in this dataset. The SLA targets here are analyst-set, "
     "clearly-labeled decision-support thresholds — adjust them on the Assumptions tab and every SLA metric updates."),
    ("REFRESH", "Re-pull data with scripts/pull_data.py, re-prepare with scripts/prepare_data.py, rebuild with "
     "scripts/build_workbook.py + finalize_workbook.py, or point Power Query at data/processed/ (see power-bi/refresh_instructions.md)."),
    ("SOURCE", "NYC Open Data — 311 Service Requests from 2010 to Present (erm2-nwe9), pulled 2026-07-05. "
     "See data-sources.md for full citation and reproduction steps."),
]
r = 5
for head, body in rows:
    if head:
        ws.cell(row=r, column=2, value=head).font = F_H2
        c = ws.cell(row=r, column=3, value=body)
        c.font = F_BODY
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = max(28, 14 * (len(body) // 95 + 1))
    r += 2

# ===========================================================================
# 2. Raw_Data_Sample
# ===========================================================================
ws = wb.create_sheet("Raw_Data_Sample")
title(ws, "Raw data sample — 2,000 rows exactly as pulled from NYC Open Data (no cleaning applied)",)
note(ws, "A2", "Shown for transparency: compare against Cleaned_Data to see what preparation changed. Full raw pull is reproducible via scripts/pull_data.py.")
hdr = list(raw_preview.columns)
for j, h in enumerate(hdr, 1):
    ws.cell(row=4, column=j, value=h)
style_header_row(ws, 4, 1, len(hdr))
for i, row in enumerate(raw_preview.itertuples(index=False), 5):
    for j, v in enumerate(row, 1):
        ws.cell(row=i, column=j, value=None if pd.isna(v) else v)
tab = Table(displayName="tblRaw", ref=f"A4:{get_column_letter(len(hdr))}{4 + len(raw_preview)}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight15", showRowStripes=True)
ws.add_table(tab)
ws.freeze_panes = "A5"
for j in range(1, len(hdr) + 1):
    ws.column_dimensions[get_column_letter(j)].width = 16

# ===========================================================================
# 3. Cleaned_Data (100K rows + derived formula columns)
# ===========================================================================
ws = wb.create_sheet("Cleaned_Data")
headers = ["Unique Key", "Created Date", "Closed Date", "Agency", "Agency Name",
           "Complaint Type", "Descriptor", "Location Type", "Incident Zip",
           "Borough", "Status", "Channel", "Slice",
           "Is Closed", "Closure Days", "Age Days", "SLA Days", "SLA Status",
           "Aging Bucket", "Created Week", "Created Month", "Age Tiebreak"]
for j, h in enumerate(headers, 1):
    ws.cell(row=1, column=j, value=h)
style_header_row(ws, 1, 1, len(headers))

data_cols = ["unique_key", "created_date", "closed_date", "agency", "agency_name",
             "complaint_type", "descriptor", "location_type", "incident_zip",
             "borough", "status", "open_data_channel_type", "slice"]
values = sample[data_cols].where(pd.notna(sample[data_cols]), None)

for i, row in enumerate(values.itertuples(index=False), 2):
    for j, v in enumerate(row, 1):
        ws.cell(row=i, column=j, value=v)
    # derived formula columns N..V
    ws.cell(row=i, column=14, value=f'=IF(AND($C{i}<>"",$C{i}>=$B{i}),1,0)')
    ws.cell(row=i, column=15, value=f'=IF($N{i}=1,$C{i}-$B{i},"")')
    ws.cell(row=i, column=16, value=f'=IF($N{i}=0,SnapshotDate-INT($B{i}),"")')
    ws.cell(row=i, column=17,
            value=f'=IFERROR(INDEX(tblSLA[SLA Days],MATCH($F{i},tblSLA[Complaint Type],0)),DefaultSLA)')
    ws.cell(row=i, column=18,
            value=f'=IF($N{i}=1,IF($O{i}<=$Q{i},"Met","Missed"),IF($P{i}<=$Q{i},"Open Within SLA","Open Past SLA"))')
    ws.cell(row=i, column=19,
            value=(f'=IF($N{i}=1,"",IF($P{i}<=AgeBucket1,"0-7 days",IF($P{i}<=AgeBucket2,"8-30 days",'
                   f'IF($P{i}<=AgeBucket3,"31-90 days",IF($P{i}<=AgeBucket4,"91-180 days","180+ days")))))'))
    ws.cell(row=i, column=20, value=f'=INT($B{i})-WEEKDAY(INT($B{i}),3)')
    ws.cell(row=i, column=21, value=f'=DATE(YEAR($B{i}),MONTH($B{i}),1)')
    ws.cell(row=i, column=22, value=f'=IF($N{i}=0,$P{i}+ROW()/10000000,"")')

# number formats (bulk, per column)
fmt_map = {2: "yyyy-mm-dd hh:mm", 3: "yyyy-mm-dd hh:mm", 15: "0.0", 16: "0",
           17: "0", 20: "yyyy-mm-dd", 21: "mmm yyyy"}
for col, fmt in fmt_map.items():
    for i in range(2, LAST + 1):
        ws.cell(row=i, column=col).number_format = fmt

tab = Table(displayName="tblClean", ref=f"A1:V{LAST}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight15", showRowStripes=False)
ws.add_table(tab)
ws.freeze_panes = "D2"
widths = [12, 17, 17, 9, 30, 24, 26, 20, 11, 13, 12, 12, 15, 9, 11, 9, 9, 15, 12, 12, 11, 11]
for j, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(j)].width = w
ws.column_dimensions["V"].hidden = True

# ===========================================================================
# 4. Data_Dictionary
# ===========================================================================
ws = wb.create_sheet("Data_Dictionary")
ws.sheet_view.showGridLines = False
title(ws, "Data dictionary")
dd = [
    ("Unique Key", "Source", "Text", "NYC 311 service-request identifier (one row per request)."),
    ("Created Date", "Source", "Datetime", "When the request was submitted to 311."),
    ("Closed Date", "Source", "Datetime", "When the responding agency closed the request (blank if still open)."),
    ("Agency", "Source", "Text", "Acronym of the responding agency (NYPD, HPD, DSNY, DOT...)."),
    ("Agency Name", "Source", "Text", "Full responding-agency name."),
    ("Complaint Type", "Source", "Text", "Request category (Illegal Parking, HEAT/HOT WATER, Noise - Residential...)."),
    ("Descriptor", "Source", "Text", "Sub-category detail beneath Complaint Type."),
    ("Location Type", "Source", "Text", "Setting of the incident (Street/Sidewalk, Residential Building...)."),
    ("Incident Zip", "Source", "Text", "ZIP code of the incident (5-digit, may be blank)."),
    ("Borough", "Source", "Text", "Borough, normalized to Title Case; blank -> 'Unspecified'."),
    ("Status", "Source", "Text", "Request lifecycle status as reported by NYC (Open, In Progress, Closed...)."),
    ("Channel", "Source", "Text", "How the request arrived (PHONE, ONLINE, MOBILE...)."),
    ("Slice", "Prepared", "Text", "'activity_q2_2026' = created in the analysis quarter; 'aged_open_backlog' = older, still open at pull time."),
    ("Is Closed", "Derived (formula)", "0/1", "1 if a valid Closed Date exists on/after Created Date."),
    ("Closure Days", "Derived (formula)", "Number", "Closed Date - Created Date, in days (closed requests only)."),
    ("Age Days", "Derived (formula)", "Number", "SnapshotDate - Created Date for OPEN requests. Ages are measured to the snapshot (2026-07-05), not 'today', so results reproduce."),
    ("SLA Days", "Derived (formula)", "Number", "Assumed SLA target for the Complaint Type, looked up from the Assumptions tab (DEFAULT fallback)."),
    ("SLA Status", "Derived (formula)", "Text", "Met / Missed (closed) or Open Within SLA / Open Past SLA (open)."),
    ("Aging Bucket", "Derived (formula)", "Text", "Open-request age bucket: 0-7 / 8-30 / 31-90 / 91-180 / 180+ days (boundaries on Assumptions)."),
    ("Created Week", "Derived (formula)", "Date", "Monday of the creation week (trend analysis)."),
    ("Created Month", "Derived (formula)", "Date", "First of the creation month."),
    ("Age Tiebreak", "Derived (formula, hidden)", "Number", "Age Days + tiny row-based offset; lets the Backlog tab rank oldest items without ties."),
]
ws["A3"], ws["B3"], ws["C3"], ws["D3"] = "Field", "Origin", "Type", "Definition"
style_header_row(ws, 3, 1, 4)
for i, row in enumerate(dd, 4):
    for j, v in enumerate(row, 1):
        c = ws.cell(row=i, column=j, value=v)
        c.font = F_BODY
        c.border = B_ALL
        c.alignment = Alignment(wrap_text=True, vertical="top")
for col, w in (("A", 18), ("B", 20), ("C", 10), ("D", 100)):
    ws.column_dimensions[col].width = w

# ===========================================================================
# 5. Assumptions
# ===========================================================================
ws = wb.create_sheet("Assumptions")
ws.sheet_view.showGridLines = False
title(ws, "Assumptions — every editable input in the workbook lives here",
      "Blue cells are inputs. Change them and all SLA, backlog and priority metrics recalculate.")

ws["B4"] = "Analysis window & snapshot"
ws["B4"].font = F_H2
lab_val = [
    ("Snapshot date (extract pulled)", "SnapshotDate", pd.Timestamp("2026-07-05"), "yyyy-mm-dd",
     "Open-request ages measured to this date. Source: pull date of scripts/pull_data.py, 2026-07-05."),
    ("Activity window start", "WindowStart", pd.Timestamp("2026-04-01"), "yyyy-mm-dd", ""),
    ("Activity window end (exclusive)", "WindowEnd", pd.Timestamp("2026-07-01"), "yyyy-mm-dd", ""),
    ("Full extract rows (population)", "PopulationRows", TOTAL_ROWS, "#,##0",
     "Source: data/processed/extract_metadata.txt (SoQL counts verified 2026-07-05)."),
    ("Workbook sample rows", "SampleRows", N, "#,##0", "Random sample, seed 42 (scripts/prepare_data.py)."),
]
r = 5
for label, name, v, fmt, src in lab_val:
    ws.cell(row=r, column=2, value=label).font = F_BODY
    c = ws.cell(row=r, column=3, value=v)
    c.font = F_INPUT
    c.number_format = fmt
    c.fill = FILL_AMBER
    wb.defined_names.add(DefinedName(name, attr_text=f"Assumptions!$C${r}"))
    if src:
        ws.cell(row=r, column=4, value=src).font = F_NOTE
    r += 1
ws.cell(row=r, column=2, value="Scaling factor (population / sample)").font = F_BODY
c = ws.cell(row=r, column=3, value="=PopulationRows/SampleRows")
c.font = F_FORMULA
c.number_format = "0.000"
wb.defined_names.add(DefinedName("ScaleFactor", attr_text=f"Assumptions!$C${r}"))
ws.cell(row=r, column=4, value="Multiply sample counts by this to estimate full-extract volumes.").font = F_NOTE

r += 2
ws.cell(row=r, column=2, value="Aging buckets & default SLA").font = F_H2
r += 1
for label, name, v, notetxt in [
    ("Aging bucket 1 upper bound (days)", "AgeBucket1", 7, "0-7 days"),
    ("Aging bucket 2 upper bound (days)", "AgeBucket2", 30, "8-30 days"),
    ("Aging bucket 3 upper bound (days)", "AgeBucket3", 90, "31-90 days"),
    ("Aging bucket 4 upper bound (days)", "AgeBucket4", 180, "91-180 days; older = 180+"),
    ("Default SLA for unlisted types (days)", "DefaultSLA", 14, "Fallback when a Complaint Type is not in the SLA table"),
]:
    ws.cell(row=r, column=2, value=label).font = F_BODY
    c = ws.cell(row=r, column=3, value=v)
    c.font = F_INPUT
    c.fill = FILL_AMBER
    wb.defined_names.add(DefinedName(name, attr_text=f"Assumptions!$C${r}"))
    ws.cell(row=r, column=4, value=notetxt).font = F_NOTE
    r += 1

r += 1
ws.cell(row=r, column=2, value="Priority model weights (must sum to 100%)").font = F_H2
r += 1
wrow = r
for label, name, v in [("Demand volume weight", "WVolume", 0.30),
                       ("Open backlog weight", "WBacklog", 0.25),
                       ("SLA miss rate weight", "WMiss", 0.25),
                       ("Aged backlog share weight", "WAged", 0.20)]:
    ws.cell(row=r, column=2, value=label).font = F_BODY
    c = ws.cell(row=r, column=3, value=v)
    c.font = F_INPUT
    c.number_format = "0%"
    c.fill = FILL_AMBER
    wb.defined_names.add(DefinedName(name, attr_text=f"Assumptions!$C${r}"))
    r += 1
ws.cell(row=r, column=2, value="Weight check").font = F_BODY
c = ws.cell(row=r, column=3, value="=WVolume+WBacklog+WMiss+WAged")
c.font = F_FORMULA
c.number_format = "0%"
ws.conditional_formatting.add(
    f"C{r}", CellIsRule(operator="notEqual", formula=["1"], font=Font(color="FFFFFF", bold=True),
                        fill=PatternFill("solid", start_color=RED_TXT)))
ws.cell(row=r, column=4, value="Turns red if the four weights do not sum to 100%.").font = F_NOTE

r += 2
ws.cell(row=r, column=2, value="SLA targets by complaint type (analyst-set decision-support thresholds — NYC publishes no per-type SLA)").font = F_H2
r += 1
sla_start = r
ws.cell(row=r, column=2, value="Complaint Type")
ws.cell(row=r, column=3, value="SLA Days")
ws.cell(row=r, column=4, value="Rationale")
style_header_row(ws, r, 2, 4)
for _, srow in sla.iterrows():
    r += 1
    ws.cell(row=r, column=2, value=srow["complaint_type"]).font = F_BODY
    c = ws.cell(row=r, column=3, value=int(srow["sla_days"]))
    c.font = F_INPUT
    ws.cell(row=r, column=4, value=srow["rationale"]).font = F_BODY
    for cc in range(2, 5):
        ws.cell(row=r, column=cc).border = B_ALL
tab = Table(displayName="tblSLA", ref=f"B{sla_start}:D{r}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight15", showRowStripes=True)
ws.add_table(tab)
for col, w in (("B", 38), ("C", 12), ("D", 60)):
    ws.column_dimensions[col].width = w

# ===========================================================================
# helper to build formula summary blocks (SLA_Analysis / Backlog_Analysis)
# ===========================================================================
K = "tblClean"


def miss_rate_formulas(crit_col, crit_ref):
    """COUNTIFS-based SLA metrics for one criterion (e.g. agency)."""
    total = f'=COUNTIFS({K}[{crit_col}],{crit_ref})'
    closed = f'=COUNTIFS({K}[{crit_col}],{crit_ref},{K}[Is Closed],1)'
    open_ = f'=COUNTIFS({K}[{crit_col}],{crit_ref},{K}[Is Closed],0)'
    missed = (f'=COUNTIFS({K}[{crit_col}],{crit_ref},{K}[SLA Status],"Missed")'
              f'+COUNTIFS({K}[{crit_col}],{crit_ref},{K}[SLA Status],"Open Past SLA")')
    known = (f'=COUNTIFS({K}[{crit_col}],{crit_ref})'
             f'-COUNTIFS({K}[{crit_col}],{crit_ref},{K}[SLA Status],"Open Within SLA")')
    avg_close = f'=IFERROR(AVERAGEIFS({K}[Closure Days],{K}[{crit_col}],{crit_ref},{K}[Is Closed],1),"")'
    return total, closed, open_, missed, known, avg_close


def build_sla_block(ws, start_row, label, crit_col, labels):
    ws.cell(row=start_row, column=1, value=label).font = F_H2
    hdr = ["", "Total (sample)", "Closed", "Open", "SLA Missed*", "SLA Outcomes Known",
           "SLA Miss Rate", "Avg Closure Days"]
    for j, h in enumerate(hdr, 1):
        ws.cell(row=start_row + 1, column=j, value=h)
    style_header_row(ws, start_row + 1, 1, len(hdr))
    r = start_row + 2
    for lab in labels:
        ref = f'$A{r}'
        ws.cell(row=r, column=1, value=lab).font = F_BODY
        t, c, o, m, k, a = miss_rate_formulas(crit_col, ref)
        for j, f in ((2, t), (3, c), (4, o), (5, m), (6, k), (8, a)):
            cell = ws.cell(row=r, column=j, value=f)
            cell.font = F_FORMULA
            cell.number_format = "#,##0" if j < 8 else "0.0"
        cell = ws.cell(row=r, column=7, value=f'=IFERROR($E{r}/$F{r},"")')
        cell.font = F_FORMULA
        cell.number_format = "0.0%"
        for j in range(1, 9):
            ws.cell(row=r, column=j).border = B_ALL
        r += 1
    ws.conditional_formatting.add(
        f"G{start_row + 2}:G{r - 1}",
        ColorScaleRule(start_type="num", start_value=0, start_color="FFFFFF",
                       end_type="num", end_value=0.6, end_color="E74C3C"))
    return r


# ===========================================================================
# 6. SLA_Analysis
# ===========================================================================
ws = wb.create_sheet("SLA_Analysis")
ws.sheet_view.showGridLines = False
title(ws, "SLA performance — met vs missed against assumed targets",
      "* SLA Missed = closed late + open past its SLA window. Miss rate excludes open requests still within SLA "
      "(outcome not yet known). Targets are assumptions — edit them on the Assumptions tab.")

ws["A4"] = "Overall (sample)"
ws["A4"].font = F_H2
kpis = [
    ("Total requests", f'=COUNTA({K}[Unique Key])', "#,##0"),
    ("SLA missed*", f'=COUNTIFS({K}[SLA Status],"Missed")+COUNTIFS({K}[SLA Status],"Open Past SLA")', "#,##0"),
    ("SLA outcomes known", f'=COUNTA({K}[Unique Key])-COUNTIFS({K}[SLA Status],"Open Within SLA")', "#,##0"),
    ("SLA miss rate", "=B6/B7", "0.0%"),
    ("Est. full-extract missed", "=B6*ScaleFactor", "#,##0"),
]
for i, (lab, f, fmt) in enumerate(kpis, 5):
    ws.cell(row=i, column=1, value=lab).font = F_BODY
    c = ws.cell(row=i, column=2, value=f)
    c.font = F_FORMULA
    c.number_format = fmt
    c.fill = FILL_KPI

r = build_sla_block(ws, 12, "By agency (top 12 by sample volume)", "Agency", top_agencies)
r = build_sla_block(ws, r + 2, "By borough", "Borough", boroughs + ["Unspecified"])
r = build_sla_block(ws, r + 2, "By complaint type (top 15 by sample volume)", "Complaint Type", top_complaints)

ws.cell(row=r + 2, column=1, value="By month created (activity window)").font = F_H2
hdr = ["", "Total (sample)", "Closed", "Open", "SLA Missed*", "SLA Outcomes Known", "SLA Miss Rate", "Avg Closure Days"]
for j, h in enumerate(hdr, 1):
    ws.cell(row=r + 3, column=j, value=h)
style_header_row(ws, r + 3, 1, len(hdr))
rr = r + 4
for month in ("2026-04-01", "2026-05-01", "2026-06-01"):
    ws.cell(row=rr, column=1, value=pd.Timestamp(month)).number_format = "mmm yyyy"
    ws.cell(row=rr, column=1).font = F_BODY
    t = f'=COUNTIFS({K}[Created Month],$A{rr})'
    cldd = f'=COUNTIFS({K}[Created Month],$A{rr},{K}[Is Closed],1)'
    op = f'=COUNTIFS({K}[Created Month],$A{rr},{K}[Is Closed],0)'
    m = (f'=COUNTIFS({K}[Created Month],$A{rr},{K}[SLA Status],"Missed")'
         f'+COUNTIFS({K}[Created Month],$A{rr},{K}[SLA Status],"Open Past SLA")')
    kk = f'=COUNTIFS({K}[Created Month],$A{rr})-COUNTIFS({K}[Created Month],$A{rr},{K}[SLA Status],"Open Within SLA")'
    a = f'=IFERROR(AVERAGEIFS({K}[Closure Days],{K}[Created Month],$A{rr},{K}[Is Closed],1),"")'
    for j, f, fmt in ((2, t, "#,##0"), (3, cldd, "#,##0"), (4, op, "#,##0"),
                      (5, m, "#,##0"), (6, kk, "#,##0"), (8, a, "0.0")):
        c = ws.cell(row=rr, column=j, value=f)
        c.font = F_FORMULA
        c.number_format = fmt
    c = ws.cell(row=rr, column=7, value=f'=IFERROR($E{rr}/$F{rr},"")')
    c.font = F_FORMULA
    c.number_format = "0.0%"
    for j in range(1, 9):
        ws.cell(row=rr, column=j).border = B_ALL
    rr += 1

ws.column_dimensions["A"].width = 34
for col in "BCDEFGH":
    ws.column_dimensions[col].width = 15
ws.freeze_panes = "A4"

# ===========================================================================
# 7. Backlog_Analysis
# ===========================================================================
ws = wb.create_sheet("Backlog_Analysis")
ws.sheet_view.showGridLines = False
title(ws, "Backlog — open requests, aging pressure and the oldest unresolved items",
      "Ages measured to the snapshot date. 'Est. full extract' = sample count x scaling factor (Assumptions).")

kpis = [
    ("Open requests (sample)", f'=COUNTIFS({K}[Is Closed],0)', "#,##0"),
    ("Share of all requests open", f'=B5/COUNTA({K}[Unique Key])', "0.0%"),
    ("Open aged 31+ days", f'=COUNTIFS({K}[Is Closed],0,{K}[Age Days],">"&AgeBucket2)', "#,##0"),
    ("Open aged 91+ days", f'=COUNTIFS({K}[Is Closed],0,{K}[Age Days],">"&AgeBucket3)', "#,##0"),
    ("Open aged 181+ days", f'=COUNTIFS({K}[Is Closed],0,{K}[Age Days],">"&AgeBucket4)', "#,##0"),
    ("Average age of open requests (days)", f'=AVERAGEIFS({K}[Age Days],{K}[Is Closed],0)', "0.0"),
    ("Est. full-extract open requests", "=B5*ScaleFactor", "#,##0"),
]
ws["A4"] = "Backlog KPIs"
ws["A4"].font = F_H2
for i, (lab, f, fmt) in enumerate(kpis, 5):
    ws.cell(row=i, column=1, value=lab).font = F_BODY
    c = ws.cell(row=i, column=2, value=f)
    c.font = F_FORMULA
    c.number_format = fmt
    c.fill = FILL_KPI

ws["A14"] = "Aging buckets x borough (open requests, sample)"
ws["A14"].font = F_H2
buckets = ["0-7 days", "8-30 days", "31-90 days", "91-180 days", "180+ days"]
for j, b in enumerate(boroughs + ["Unspecified", "TOTAL"], 2):
    ws.cell(row=15, column=j, value=b)
ws.cell(row=15, column=1, value="Aging bucket")
style_header_row(ws, 15, 1, 9)
for i, b in enumerate(buckets, 16):
    ws.cell(row=i, column=1, value=b).font = F_BODY
    for j, bor in enumerate(boroughs + ["Unspecified"], 2):
        col_letter = get_column_letter(j)
        f = f'=COUNTIFS({K}[Aging Bucket],$A{i},{K}[Borough],{col_letter}$15)'
        c = ws.cell(row=i, column=j, value=f)
        c.font = F_FORMULA
        c.number_format = "#,##0"
        c.border = B_ALL
    c = ws.cell(row=i, column=8, value=f"=SUM(B{i}:G{i})")
    c.font = F_FORMULA
    c.number_format = "#,##0"
    c.border = B_ALL
ws.conditional_formatting.add(
    "B16:G20", DataBarRule(start_type="num", start_value=0, end_type="max",
                           color=TEAL, showValue=True))

ws["A23"] = "Weekly flow — created vs closed-by-snapshot (activity window, sample)"
ws["A23"].font = F_H2
for j, h in enumerate(["Week starting", "Created", "Closed by snapshot", "Still open"], 1):
    ws.cell(row=24, column=j, value=h)
style_header_row(ws, 24, 1, 4)
week_starts = pd.date_range("2026-03-30", "2026-06-29", freq="W-MON")
for i, wk in enumerate(week_starts, 25):
    ws.cell(row=i, column=1, value=wk).number_format = "yyyy-mm-dd"
    ws.cell(row=i, column=1).font = F_BODY
    for j, f in ((2, f'=COUNTIFS({K}[Created Week],$A{i})'),
                 (3, f'=COUNTIFS({K}[Created Week],$A{i},{K}[Is Closed],1)'),
                 (4, f'=COUNTIFS({K}[Created Week],$A{i},{K}[Is Closed],0)')):
        c = ws.cell(row=i, column=j, value=f)
        c.font = F_FORMULA
        c.number_format = "#,##0"
        c.border = B_ALL
week_last = 24 + len(week_starts)

ws["F23"] = "High-risk unresolved: 20 oldest open requests (sample)"
ws["F23"].font = F_H2
for j, h in enumerate(["Rank", "Age (days)", "Unique Key", "Complaint Type", "Borough", "Agency"], 6):
    ws.cell(row=24, column=j, value=h)
style_header_row(ws, 24, 6, 11)
for i in range(25, 45):
    k = i - 24
    ws.cell(row=i, column=6, value=k).font = F_BODY
    c = ws.cell(row=i, column=7, value=f'=INT(LARGE({K}[Age Tiebreak],{k}))')
    c.font = F_FORMULA
    c.number_format = "#,##0"
    for j, colname in ((8, "Unique Key"), (9, "Complaint Type"), (10, "Borough"), (11, "Agency")):
        f = (f'=INDEX({K}[{colname}],MATCH(LARGE({K}[Age Tiebreak],{k}),{K}[Age Tiebreak],0))')
        c = ws.cell(row=i, column=j, value=f)
        c.font = F_FORMULA
    for j in range(6, 12):
        ws.cell(row=i, column=j).border = B_ALL

ws.column_dimensions["A"].width = 38
for col in "BCDE":
    ws.column_dimensions[col].width = 14
ws.column_dimensions["F"].width = 7
ws.column_dimensions["G"].width = 11
ws.column_dimensions["H"].width = 13
for col in ("I", "J", "K"):
    ws.column_dimensions[col].width = 24

# native line chart for weekly flow
from openpyxl.chart import LineChart, Reference
chart = LineChart()
chart.title = "Weekly created vs closed (sample)"
chart.height = 7
chart.width = 16
data = Reference(ws, min_col=2, max_col=3, min_row=24, max_row=week_last)
cats = Reference(ws, min_col=1, min_row=25, max_row=week_last)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.style = 2
ws.add_chart(chart, "A40")

# ===========================================================================
# 8. Priority_Model
# ===========================================================================
ws = wb.create_sheet("Priority_Model")
ws.sheet_view.showGridLines = False
title(ws, "Priority model — where should operational attention go first?",
      "Transparent weighted score per complaint-type x borough cell. Weights are on the Assumptions tab; "
      "change them and the ranking reorders. Score = volume percentile x WVolume + open-backlog percentile x WBacklog "
      "+ SLA miss rate x WMiss + aged-share-of-open x WAged.")

hdr = ["Complaint Type", "Borough", "Total (sample)", "Open", "SLA Missed*", "SLA Outcomes Known",
       "SLA Miss Rate", "Aged 31+ Open", "Aged Share of Open", "Volume Percentile",
       "Backlog Percentile", "Priority Score", "Rank"]
for j, h in enumerate(hdr, 1):
    ws.cell(row=4, column=j, value=h)
style_header_row(ws, 4, 1, len(hdr))

r = 5
first = r
for ct in top_complaints:
    for bor in boroughs:
        ws.cell(row=r, column=1, value=ct).font = F_BODY
        ws.cell(row=r, column=2, value=bor).font = F_BODY
        base = f'{K}[Complaint Type],$A{r},{K}[Borough],$B{r}'
        cells = {
            3: (f'=COUNTIFS({base})', "#,##0"),
            4: (f'=COUNTIFS({base},{K}[Is Closed],0)', "#,##0"),
            5: (f'=COUNTIFS({base},{K}[SLA Status],"Missed")+COUNTIFS({base},{K}[SLA Status],"Open Past SLA")', "#,##0"),
            6: (f'=COUNTIFS({base})-COUNTIFS({base},{K}[SLA Status],"Open Within SLA")', "#,##0"),
            7: (f'=IFERROR($E{r}/$F{r},0)', "0.0%"),
            8: (f'=COUNTIFS({base},{K}[Is Closed],0,{K}[Age Days],">"&AgeBucket2)', "#,##0"),
            9: (f'=IFERROR($H{r}/$D{r},0)', "0.0%"),
        }
        for j, (f, fmt) in cells.items():
            c = ws.cell(row=r, column=j, value=f)
            c.font = F_FORMULA
            c.number_format = fmt
        r += 1
last = r - 1
for i in range(first, last + 1):
    c = ws.cell(row=i, column=10, value=f'=PERCENTRANK.INC($C${first}:$C${last},$C{i})')
    c.font = F_FORMULA
    c.number_format = "0%"
    c = ws.cell(row=i, column=11, value=f'=PERCENTRANK.INC($D${first}:$D${last},$D{i})')
    c.font = F_FORMULA
    c.number_format = "0%"
    c = ws.cell(row=i, column=12,
                value=f'=WVolume*$J{i}+WBacklog*$K{i}+WMiss*$G{i}+WAged*$I{i}')
    c.font = F_FORMULA
    c.number_format = "0.000"
    c = ws.cell(row=i, column=13, value=f'=RANK($L{i},$L${first}:$L${last})')
    c.font = F_FORMULA
    c.number_format = "0"
    for j in range(1, 14):
        ws.cell(row=i, column=j).border = B_ALL

ws.conditional_formatting.add(
    f"L{first}:L{last}",
    ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color="F4A259"))
ws.conditional_formatting.add(
    f"M{first}:M{last}",
    CellIsRule(operator="lessThanOrEqual", formula=["10"],
               font=Font(bold=True, color="FFFFFF"), fill=PatternFill("solid", start_color=TEAL)))
note(ws, f"A{last + 2}",
     "* SLA Missed = closed late + open past SLA. Grid covers the top 15 complaint types x 5 boroughs; "
     "cells with very low volume naturally rank low. The SQL layer (sql/kpi_views.sql) scores every cell with >= 200 requests as a cross-check.")
ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 14
for j in range(3, 14):
    ws.column_dimensions[get_column_letter(j)].width = 13
ws.freeze_panes = "C5"

# ===========================================================================
# 9. Action_List
# ===========================================================================
ws = wb.create_sheet("Action_List")
ws.sheet_view.showGridLines = False
title(ws, "Action list — top operational priorities",
      "Ranked live from the Priority_Model tab (green values are cross-sheet links; they reorder if you change weights).")

ws["A4"] = "Headline KPIs"
ws["A4"].font = F_H2
for i, (lab, f, fmt) in enumerate([
    ("Total requests (sample)", "=COUNTA(tblClean[Unique Key])", "#,##0"),
    ("Open requests (sample)", "=Backlog_Analysis!B5", "#,##0"),
    ("SLA miss rate", "=SLA_Analysis!B8", "0.0%"),
    ("Open aged 31+ days", "=Backlog_Analysis!B7", "#,##0"),
], 5):
    ws.cell(row=i, column=1, value=lab).font = F_BODY
    c = ws.cell(row=i, column=2, value=f)
    c.font = F_LINK
    c.number_format = fmt
    c.fill = FILL_KPI

ws["A11"] = "Top 10 priority hotspots (live from Priority_Model)"
ws["A11"].font = F_H2
for j, h in enumerate(["Rank", "Complaint Type", "Borough", "Priority Score",
                       "Open (sample)", "SLA Miss Rate"], 1):
    ws.cell(row=12, column=j, value=h)
style_header_row(ws, 12, 1, 6)
PM_FIRST, PM_LAST = 5, 79
for i in range(13, 23):
    k = i - 12
    ws.cell(row=i, column=1, value=k).font = F_BODY
    m = f'MATCH({k},Priority_Model!$M${PM_FIRST}:$M${PM_LAST},0)'
    for j, colrange, fmt in ((2, "$A", "@"), (3, "$B", "@"), (4, "$L", "0.000"),
                             (5, "$D", "#,##0"), (6, "$G", "0.0%")):
        f = f'=INDEX(Priority_Model!{colrange}${PM_FIRST}:{colrange}${PM_LAST},{m})'
        c = ws.cell(row=i, column=j, value=f)
        c.font = F_LINK
        c.number_format = fmt
    for j in range(1, 7):
        ws.cell(row=i, column=j).border = B_ALL

ws["A25"] = "Recommended focus areas (analyst interpretation of the current ranking)"
ws["A25"].font = F_H2
recs = [
    ("Parks-related backlog (DPR) in Brooklyn and Queens",
     "Tree and parks-facility requests (Overgrown Tree/Branches, Root/Sewer/Sidewalk Condition, Maintenance or "
     "Facility) dominate the hotspot ranking: high volume, 65%+ of requests left open, and most open items already "
     "aged past 30 days. Recommend a dedicated aging review for DPR work orders with close-or-escalate triage for "
     "everything past 90 days."),
    ("TLC complaint case backlog",
     "Taxi and For-Hire-Vehicle complaints show ~86-99% SLA miss with thousands of open cases in Manhattan, Queens "
     "and Brooklyn. These are case-review workflows, not field repairs — a process/throughput review is the lever, "
     "and the assumed 30-day SLA should be validated against actual case-handling policy."),
    ("Helicopter-noise requests: reporting artifact to resolve first",
     "All ~19,000 EDC helicopter-noise requests in the window are unclosed, which reads as a workflow/recording "
     "artifact (complaints logged but never resolved in 311) rather than pure operational failure. Confirm how EDC "
     "records outcomes before treating this as the #1 operational problem — but do not let it silently distort "
     "citywide backlog KPIs; consider reporting it as its own line."),
    ("Seasonal watch: HEAT/HOT WATER in the Bronx",
     "Even in Q2, Bronx heat/hot-water requests rank in the top 15 hotspots on volume with 99% of open items aged "
     "31+ days. This becomes the dominant winter pressure — clear the aged residue now, before the October heat "
     "season begins."),
    ("Data hygiene",
     "Requests with 'Unspecified' borough (~0.15%) can't be routed geographically and are excluded from hotspot "
     "scoring. Track their share monthly (SQL check #07); a rising share degrades every location view."),
]
r = 26
for head, body in recs:
    ws.cell(row=r, column=1, value="• " + head).font = Font(name=FONT, size=11, bold=True, color=DARK)
    r += 1
    c = ws.cell(row=r, column=1, value=body)
    c.font = F_BODY
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 40
    r += 2

ws.cell(row=r, column=1, value="Suggested operating rhythm").font = F_H2
r += 1
for line in [
    "WEEKLY — refresh extract; review SLA miss rate, weekly created-vs-closed flow, and any hotspot entering the top 10.",
    "MONTHLY — trend review: complaint mix shift, borough pressure, aging distribution; re-baseline SLA targets if misses are structural.",
    "ESCALATE WHEN — a hotspot's SLA miss rate exceeds 60%, or its 31+ day open count grows two weeks in a row, or any item crosses 180 days.",
]:
    c = ws.cell(row=r, column=1, value=line)
    c.font = F_BODY
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 28
    r += 1

ws.column_dimensions["A"].width = 40
for col in "BCDEF":
    ws.column_dimensions[col].width = 16

# ===========================================================================
# 10. Pivot_Analysis placeholder (pivots added by finalize_workbook.py)
# ===========================================================================
ws = wb.create_sheet("Pivot_Analysis", 5)
ws.sheet_view.showGridLines = False
title(ws, "Pivot analysis — demand cuts",
      "Native Excel PivotTables over the Cleaned_Data table. Right-click > Refresh after editing data. "
      "Built by scripts/finalize_workbook.py.")

wb.save(OUT)
print(f"saved {OUT} ({N:,} data rows)")
